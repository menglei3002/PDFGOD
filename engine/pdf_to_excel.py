"""PDF to Excel conversion using tabula-py and camelot-py."""

import os


def convert(input_path: str, output_path: str, report_fn) -> str:
    report_fn(10, "Analyzing PDF for tables...")

    errors = []
    tables_found = False

    # Try camelot first (better accuracy for bordered tables)
    try:
        import camelot
        report_fn(15, "Trying camelot for table extraction...")
        tables = camelot.read_pdf(input_path, pages="all", flavor="lattice")
        if len(tables) > 0:
            report_fn(30, f"Camelot found {len(tables)} tables")
            import openpyxl
            wb = openpyxl.Workbook()
            for i, table in enumerate(tables):
                df = table.df
                pct = 30 + int(50 * (i + 1) / len(tables))
                report_fn(pct, f"Processing table {i+1}/{len(tables)}...")
                ws = wb.active if i == 0 else wb.create_sheet(f"Table_{i+1}")
                for r_idx, row in df.iterrows():
                    for c_idx, val in enumerate(row):
                        ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
            wb.save(output_path)
            tables_found = True
    except ImportError:
        errors.append("camelot not installed")
    except Exception as e:
        errors.append(f"camelot: {e}")

    # Fallback to tabula-py
    if not tables_found:
        try:
            import tabula
            report_fn(20, "Using tabula for table extraction...")
            dfs = tabula.read_pdf(input_path, pages="all", multiple_tables=True)

            if dfs and len(dfs) > 0:
                report_fn(40, f"Tabula found {len(dfs)} tables")
                import openpyxl
                wb = openpyxl.Workbook()
                for i, df in enumerate(dfs):
                    pct = 40 + int(50 * (i + 1) / len(dfs))
                    report_fn(pct, f"Processing table {i+1}/{len(dfs)}...")
                    ws = wb.active if i == 0 else wb.create_sheet(f"Table_{i+1}")
                    for r_idx, row in df.iterrows():
                        for c_idx, val in enumerate(row):
                            ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
                wb.save(output_path)
                tables_found = True
        except ImportError:
            errors.append("tabula not installed (needs Java)")
        except Exception as e:
            errors.append(f"tabula: {e}")

    if not tables_found:
        if errors:
            raise RuntimeError(
                "Table extraction unavailable. Both engines failed:\n- "
                + "\n- ".join(errors)
                + "\n\nTo fix: install Java (for tabula) or Ghostscript (for camelot)."
            )
        # If no engine errors but no tables found, use openpyxl directly with info
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "No Tables"
        ws.cell(row=1, column=1, value="No tables detected in this PDF.")
        ws.cell(row=2, column=1, value="Try converting to TXT or Word instead.")
        wb.save(output_path)

    if os.path.exists(output_path):
        return os.path.abspath(output_path)
    raise RuntimeError("Excel conversion failed")
